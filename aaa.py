from lxml import html
from openpyxl import Workbook
import re
import requests

class zhongyao():
    def __init__(self):
        self.text_all = dict()
        self.url="https://www.zhzyw.com/zyts/pfmf/Index.html"
        self.basic_url="https://www.zhzyw.com/"
        self.headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36"}
    def get_parse_html(self,url): #将网页源代码转换成xpath对象的函数
        res=requests.get(url,headers=self.headers)
        res.encoding = "gb2312"
        html_text = res.text
        parse_html = html.etree.HTML(html_text)#将网页源代码转换成xpath对象
        return parse_html
    def get_one(self,url):#拼接并得到xpath解析的网页
        text_url1one = self.basic_url + ''.join(url)
        en = self.get_parse_html(text_url1one)
        return en
    def get_url(self):
        en = self.get_parse_html(self.url)
        text_url1 = en.xpath('//*[@id="title"]/ul/li[1]/a/@href')
        text_url2 = en.xpath('//*[@id="title"]/ul/li[2]/a/@href')
        text_url3 = en.xpath('//*[@id="title"]/ul/li[3]/a/@href')
        text_url4 = en.xpath('//*[@id="title"]/ul/li[4]/a/@href')
        text_url5 = en.xpath('//*[@id="title"]/ul/li[5]/a/@href')
        text_url6 = en.xpath('//*[@id="title"]/ul/li[6]/a/@href')
        text_url7 = en.xpath('//*[@id="title"]/ul/li[7]/a/@href')
        enone = self.get_one(text_url7)
        text_urltwo = enone.xpath('//*[@id="left"]/div[4]/ul/li/a/@href')  # 获取某一科下面所有秘方
        self.get_nate(text_urltwo)


    def get_text_part(self,parse_html_text):
        try:
            text_alias = parse_html_text.xpath('//*[@id="wzdh"]/a[5]/text()')
            text_alias = ''.join(text_alias)
        except:
            return "没找到"
        return text_alias

    def get_text_drug(self, parse_html_text):
        try:

            text_smell = parse_html_text.xpath('//*[@id="left"]/h1/text()')
            text_smell = ''.join(text_smell)

        except:
            return "找不到"
        return text_smell
    def get_text_cure(self, parse_html_text):
        try:
            text_cure1 = parse_html_text.xpath('//*[@id="left"]/div[2]/text()')
            text_cure1 = ''.join(text_cure1)
        except:
            return "找不到"
        return text_cure1
    def get_nate(self,text_url):
        count = 0
        rows = []
        for link in text_url:#对所有帖子的站内链接进行遍历 拼接完整的帖子链接
            t_url=self.basic_url+link#拼接得到帖子的url
            parse_html_text=self.get_parse_html(t_url)
            text_name = parse_html_text.xpath('//*[@id="left"]/div[3]/ul/li/a/@href')
            for t_link in text_name:
                all_url=self.basic_url+t_link
                all_html_text=self.get_parse_html(all_url)
                text_part=self.get_text_part(all_html_text)
                text_drug=self.get_text_drug(all_html_text)
                text_prescript=self.get_text_cure(all_html_text)
                #print(text_part+"\n"+text_drug+"\n"+text_prescript)
                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'drug'
                ws['B1'] = 'prescript'
                ws['C1'] = 'part'

                row=[text_drug,text_prescript,text_part]
                rows.append(row)
                for new_row in rows:
                    ws.append(new_row)
                count+=1
                print(count)
                wb.save(text_part+'.xlsx')


zhongyao = zhongyao()
zhongyao.get_url()
